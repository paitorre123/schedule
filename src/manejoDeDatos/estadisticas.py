import openpyxl
from src.manejoDeDatos import grafico
from src.manejoDeDatos.ManejoDeDatos import ManejoDeDatos
from collections import OrderedDict
import numpy as np
import os
from decimal import *
from src.manejoDeDatos.grafico import Grafico
import itertools

class Estadisticas(object):
    URL_EXCEL = '/'
    def __init__(self, numeroPoi, numeroUsuarios, anonimato, efectividad, talla, tecnica):
        '''
        estos datos son utilizados para crear estadisticas
        :param numeroPoi:  numero de puntos de interes
        :param numeroUsuarios:  numero de usuarios
        :param anonimato: k anonimato de las pruebas
        '''
        self.numeroUsuarios = numeroUsuarios
        self.numeroPoi = numeroPoi
        self.anonimato = anonimato
        self.tecnica = tecnica
        self.efectividad = efectividad
        self.talla = talla
        #self.numero_cues_arriban_en_el_servidor()
        self.numero_cues_creadas()
        self.cues_creadas_respondidas()
        #self.cues_recibidas_respondidas()
        #self.tiempo_cues_creadas()
        #self.tiempo_cues_recibidas()
        #self.tiempo_servidor()
        #self.talla_cues_creadas()
        #self.cues_creadas_por_usuarios()
        #self.cues_enviadas_por_usuarios()
        #self.cues_creadas_obsoletas()
        #self.cues_creadas_obsoletas_enviadas()
        #self.elementos_usados_por_cronograma()
        #self.razon_de_rendimiento()
        #self.utilidad_y_efectividad_del_programa()
        #self.justicia()

    def numero_cues_arriban_en_el_servidor(self):
        print ':::::::::::::::::::::::::::::::'
        #1. datos para recorrer la hoja excel
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_SERVER_CUES)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        #2. atributos para almacenar las cues que arriban
        self.cuesQueArriban = OrderedDict()
        #3. recorrido de la hoja exel
        for r in range(2, max_row):
            #print sum([len(x) for x in self.cuesQueArriban.values()])
            self.cuesQueArriban[hojaDeTrabajo.cell(row=r, column=1).value] = []
            for c in range(2, max_column):
                if hojaDeTrabajo.cell(row=r, column=c).value != None:
                    self.cuesQueArriban[hojaDeTrabajo.cell(row=r, column=1).value].append(hojaDeTrabajo.cell(row=r, column=c).value)

        #4. mostrar datos de la hoja
        self.maximoTiempoDeCreacion = 0
        for key, value in self.cuesQueArriban.items():
            print 'tiempo: {} -> Cues que arriban: {}'.format(key, len(value))
            for cues in value:
                if self.maximoTiempoDeCreacion < int(str(cues).split(':')[1]):
                    self.maximoTiempoDeCreacion = int(str(cues).split(':')[1])
        self.tiempoDemandaCumplida = max(self.cuesQueArriban)
        print 'Tiempo demanda cumplida: {}'.format(max(self.cuesQueArriban))
        print 'Demanda total: {}'.format(sum([len(x) for x in self.cuesQueArriban.values()]))
        print 'Maximo tiempo de creacion: {}'.format(self.maximoTiempoDeCreacion)


    def numero_cues_creadas(self):
        print ':::::::::::::::::::::::::::::::'
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_USER_CREATED_CUES)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        numeroCuesTotales = 0
        cuesCreadasEnTiempoQueSeCumpleLaDemanda = []
        for r in range(1, max_row):
            for c in range(2, max_column):
                if hojaDeTrabajo.cell(row=r, column=c).value != None:
                    valor = hojaDeTrabajo.cell(row=r, column=c).value
                    datos = valor.split(":")
                    numeroCuesTotales += 1
                    cuesCreadasEnTiempoQueSeCumpleLaDemanda.append((valor, datos[1]))

        #-----------------
        res = OrderedDict()
        for v, k in cuesCreadasEnTiempoQueSeCumpleLaDemanda:
            if k in res:
                res[k].append(v)
            else:
                res[k] = [v]
        [{'type': k, 'items': v} for k, v in res.items()]
        self.cuesTotalesCreadas = 0
        # orden de forma ascendente el diccionario
        res = OrderedDict(sorted(res.items(), key=lambda x: int(x[0])))
        for key, value in res.items():
            print 'Tiempo:{} -> Cues creadas:{}'.format(key, len(value))
            self.cuesTotalesCreadas += len(value)
            self.tiempoCreacionUltimasCues = key
        print 'CUEs totales creadas: {}'.format(self.cuesTotalesCreadas)
        # -------------------------------------
        # variable de retorno de la informacion
        # -------------------------------------
        self.cuesCreadasPorTiempo = res

    def cues_creadas_respondidas(self):
        print ':::::::::::::::::::::::::::::::'
        # 1. atributo para almacenar el tiempo-1 en que se reciben los elementos exactos de las cues
        self.tiempoElementosCreados = OrderedDict()
        for key, value in self.cuesCreadasPorTiempo.items():
            for cue in value:
                self.tiempoElementosCreados[cue] = OrderedDict()

        # 2. datos para recorrer la hoja excel
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_CUES_USAN_ELEMENTOS)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        time = 1
        inicio = 1
        termino = 0
        self.cuesCreadasRespondidas = OrderedDict()
        self.programas = OrderedDict()
        for r in range(1, max_row):
            if hojaDeTrabajo.cell(row=r, column=1).value != None:
                #print 'Time: ', time,
                for c in range(2, max_column):
                    if hojaDeTrabajo.cell(row=r, column=c).value != None:
                        if hojaDeTrabajo.cell(row=r, column=c).value in self.tiempoElementosCreados:
                            #print '[', hojaDeTrabajo.cell(row=r, column=c).value, ':',  hojaDeTrabajo.cell(row=r, column=1).value, ']',
                            self.tiempoElementosCreados[hojaDeTrabajo.cell(row=r, column=c).value][ hojaDeTrabajo.cell(row=r, column=1).value] = [time]
                #print ''
                time += 1
            else:
                if hojaDeTrabajo.cell(row=r - 1, column=1).value != None:
                    time += 1
                    termino = time - 1
                    k = '{}:{}'.format(inicio, termino)
                    self.cuesCreadasRespondidas[k] = []
                    self.programas[k] = []
                    inicio = time
            if r == max_row-1:
                time += 1
                termino = time - 1
                k = '{}:{}'.format(inicio, termino)
                self.cuesCreadasRespondidas[k] = []
                self.programas[k] = []
        # dejamos solo los elementos de datos que son respondidos
        self.tiempoElementosCreados = {k: v for k, v in self.tiempoElementosCreados.items() if len(v)>0}

        for key, values in self.tiempoElementosCreados.items():
            consultaEn = ''
            for lapso in self.cuesCreadasRespondidas:
                menor = str(lapso).split(':')[0]
                mayor = str(lapso).split(':')[1]
                if max(list(values.values()))[0] <= int(mayor) and max(list(values.values()))[0] >= int(menor):
                    consultaEn =  lapso
            self.cuesCreadasRespondidas[consultaEn].append(key)
        # Dejando los cronogramas que responden determinadas consultas
        empty_keys = [k for k, v in self.cuesCreadasRespondidas.iteritems() if not v]
        for k in empty_keys:
            del self.cuesCreadasRespondidas[k]
        for key, value in self.cuesCreadasRespondidas.items():
            print '{} -> cues creadas respondidas: {}'.format(key, len(value))
        print 'Total Consultas respondidas: {}'.format(sum([len(x) for x in self.cuesCreadasRespondidas.values()]))
        for key in self.cuesCreadasRespondidas:
            self.tiempoDemandaCumplidaCuesCreadas = int(str(key).split(':')[1])
        print 'Tiempo demanda cumplida: {}'.format(self.tiempoDemandaCumplidaCuesCreadas)

    def cues_recibidas_respondidas(self):
        print ':::::::::::::::::::::::::::::::'
        # 1. atributo para almacenar el tiempo-1 en que se reciben los elementos exactos de las cues
        self.tiempoElementosRecibidos = OrderedDict()
        for key, value in self.cuesQueArriban.items():
            for cue in value:
                self.tiempoElementosRecibidos[cue] = OrderedDict()

        # 2. datos para recorrer la hoja excel
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_CUES_USAN_ELEMENTOS)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        time = 1
        inicio = 1
        termino = 0
        self.cuesRecibidasRespondidas = OrderedDict()
        self.elementosPorPrograma = OrderedDict()
        for r in range(1, max_row):
            if hojaDeTrabajo.cell(row=r, column=1).value != None:
                # print 'Time: ', time,
                for c in range(2, max_column):
                    if hojaDeTrabajo.cell(row=r, column=c).value != None:
                        if hojaDeTrabajo.cell(row=r, column=c).value in self.tiempoElementosRecibidos:
                            # print '[', hojaDeTrabajo.cell(row=r, column=c).value, ':',  hojaDeTrabajo.cell(row=r, column=1).value, ']',
                            self.tiempoElementosRecibidos[hojaDeTrabajo.cell(row=r, column=c).value][hojaDeTrabajo.cell(row=r, column=1).value] = [time]
                # print ''
                time += 1
            else:
                if hojaDeTrabajo.cell(row=r - 1, column=1).value != None:
                    time += 1
                    termino = time - 1
                    k = '{}:{}'.format(inicio, termino)
                    self.cuesRecibidasRespondidas[k] = []
                    self.elementosPorPrograma[k] = []
                    inicio = time
            if r == max_row - 1:
                time += 1
                termino = time - 1
                k = '{}:{}'.format(inicio, termino)
                self.cuesRecibidasRespondidas[k] = []
                self.elementosPorPrograma[k] = []
        # dejamos solo los elementos de datos que son respondidos
        self.tiempoElementosRecibidos = {k: v for k, v in self.tiempoElementosRecibidos.items() if len(v) > 0}

        for key, values in self.tiempoElementosRecibidos.items():
            consultaEn = ''
            for lapso in self.cuesRecibidasRespondidas:
                menor = str(lapso).split(':')[0]
                mayor = str(lapso).split(':')[1]
                if max(list(values.values()))[0] <= int(mayor) and max(list(values.values()))[0] >= int(menor):
                    consultaEn = lapso
            self.cuesRecibidasRespondidas[consultaEn].append(key)
        # Dejando los cronogramas que responden determinadas consultas
        empty_keys = [k for k, v in self.cuesRecibidasRespondidas.iteritems() if not v]
        for k in empty_keys:
            del self.cuesRecibidasRespondidas[k]
        for key, value in self.cuesRecibidasRespondidas.items():
            print '{} -> cues  recibidas respondidas: {}'.format(key, len(value))
        print 'Total Consultas recibidas respondidas: {}'.format(sum([len(x) for x in self.cuesRecibidasRespondidas.values()]))
        #self.elementos_por_programa()

    def elementos_por_programa(self):
        print ':::::::::::::::::::::::::::::::'
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_CUES_USAN_ELEMENTOS)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        time = 1
        inicio = 1
        termino = 0

        for r in range(1, max_row):
            if hojaDeTrabajo.cell(row=r, column=1).value != None:
                # print 'Time: ', time,

                for lapso in self.elementosPorPrograma:
                    inferior = int(str(lapso).split(':')[0])
                    superior = int(str(lapso).split(':')[1])
                    if time < superior and time >= inferior:
                        self.elementosPorPrograma[lapso].append(hojaDeTrabajo.cell(row=r, column=1).value)
                # print ''
                time += 1
            else:
                if hojaDeTrabajo.cell(row=r - 1, column=1).value != None:
                    time += 1
                    termino = time - 1
                    inicio = time
            if r == max_row - 1:
                time += 1
                termino = time - 1
        self.valor_seleccion_por_programa()
        if '1:{}'.format(self.talla+2) in self.elementosPorPrograma:
            del self.elementosPorPrograma['1:{}'.format(self.talla+2)]
        if '1:{}'.format(self.talla+1) in self.elementosPorPrograma:
            del self.elementosPorPrograma['1:{}'.format(self.talla+1)]

        for key, value in self.elementosPorPrograma.items():
            if len(self.valorSeleccion) > 0:
                try:
                    print '{}-> Elementos por programa:{}--> valor de seleccion: {}'.format(key, len(value),self.valorSeleccion[len(value)] )
                except KeyError:
                    del self.elementosPorPrograma[key]
                    print KeyError
                    os.system('pause')
            else:
                print '{}-> Elementos por programa:{}'.format(key, len(value))
        #os.system('pause')

    def valor_seleccion_por_programa(self):
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_VALORES_SELECCION)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        self.valorSeleccion = OrderedDict()

        for r in range(2, max_row):
            self.valorSeleccion[hojaDeTrabajo.cell(row=r, column=3).value] = hojaDeTrabajo.cell(row=r, column=2).value




    def subconsultas_respondidas(self):
        '''
        Las sub consultas respondidas pertenecen a las
        cues que son enviadas al servidor, ya que antes
        de ser enviadas son solo consultas de rango que
        no contienen sub consultas.
        :return: subconsultas que son respondidas por tiempo
        '''
        print ':::::::::::::::::::::::::::::::'

        self.subconsultasDeCues = OrderedDict()

        # x. datos para recorrer la hoja excel
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_ITEM_SUBCONSULTAS)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        # x. recorrido de la hoja excel
        for r in range(1, max_row):
            cue = str(hojaDeTrabajo.cell(row=r, column=1).value).split('-')[1]
            subconsulta = str(hojaDeTrabajo.cell(row=r, column=1).value)
            if cue in self.subconsultasDeCues:
                self.subconsultasDeCues[cue][subconsulta] = []
                for c in range(2, max_column):
                    if hojaDeTrabajo.cell(row=r, column=c).value != None:
                        # se ingresan los elementos requeridos de cada sub consulta
                        self.subconsultasDeCues[cue][subconsulta].append(hojaDeTrabajo.cell(row=r, column=c).value)
            else:
                self.subconsultasDeCues[cue] = OrderedDict()
                self.subconsultasDeCues[cue][subconsulta] = []
                for c in range(2, max_column):
                    if hojaDeTrabajo.cell(row=r, column=c).value != None:
                        # se ingresan los elementos requeridos de cada sub consulta
                        self.subconsultasDeCues[cue][subconsulta].append(hojaDeTrabajo.cell(row=r, column=c).value)


        # x. Eliminacion de las cues que no son enviadas al servidor
        self.subconsultasDeCues = {k:v for k, v in self.subconsultasDeCues.items() if k in set(itertools.chain(*self.cuesQueArriban.values()))}

        #
        print 'numero cues: {}'.format(len(self.subconsultasDeCues))



        self.subconsultasEnviadasRespondidas = OrderedDict()
        for lapso in self.cuesRecibidasRespondidas: self.subconsultasEnviadasRespondidas[lapso] = OrderedDict()

        #self.cuesRecibidasRespondidas

        for lapso, cues in self.cuesRecibidasRespondidas.items():
            for cue in cues:
                print '{}'.format(cue)
                for subcues in self.subconsultasDeCues[cue]:
                    print '{}'.format(subcues)
                    print '{}'.format(self.subconsultasDeCues[cue][subcues])
                    self.subconsultasDeCues[cue][subcues] = list(set(self.subconsultasDeCues[cue][subcues])-set(self.elementosPorPrograma[lapso]))
                    print '{}'.format(self.subconsultasDeCues[cue][subcues])
                    os.system('pause')
                    self.subconsultasDeCues[cue][subcues]
                self.subconsultasEnviadasRespondidas[lapso][cue] = self.subconsultasDeCues[cue].values()



        for lapso, cues in self.subconsultasEnviadasRespondidas.items():
            print '{}->{}'.format(lapso, cues)





        os.system('pause')

    def tiempo_cues_creadas(self):
        print ':::::::::::::::::::::::::::::::'
        #1. tiempo-1 de espera de las consultas
        self.tiempoEsperaConsultas = OrderedDict()
        for key, values in self.tiempoElementosCreados.items():
            tiempoCreacion = str(key).split(':')[1]
            self.tiempoEsperaConsultas[key] = [max(list(values.values()))[0] - int(tiempoCreacion)]

            #print '{}->  tiempo-1 espera: {}'.format(key, max(list(values.values()))[0] - int(tiempoCreacion))
        print 'Promedio Tiempo de espera de las consultas creadas: {}'.format(np.average(list(self.tiempoEsperaConsultas.values())))
        print 'Varianza Tiempo de espera de las consultas creadas: {}'.format(np.var(list(self.tiempoEsperaConsultas.values())))
        print 'Desviacion estandar de espera de las consultas creadas: {}'.format(np.std(list(self.tiempoEsperaConsultas.values())))
        #2. Tiempo de servicio de las consultas
        self.tiempoServicioConsultas = OrderedDict()
        for key, values in self.tiempoElementosCreados.items():
            self.tiempoServicioConsultas[key] = [max(list(values.values()))[0] - min(list(values.values()))[0]]
        print 'Promedio Tiempo de servicio de las consultas creadas: {}'.format(np.average(list(self.tiempoServicioConsultas.values())))
        print 'Varianza Tiempo de servicio de las consultas creadas: {}'.format(np.var(list(self.tiempoServicioConsultas.values())))
        print 'Desviacion estandar de servicio de las consultas creadas: {}'.format(np.std(list(self.tiempoServicioConsultas.values())))
        #3. tiempo-1 de respuesta de las consultas
        self.tiempoRespuestaConsultas = OrderedDict()
        for key, values in self.tiempoElementosCreados.items():
            tiempoEnvio = str(key).split(':')[2]
            self.tiempoRespuestaConsultas[key] = [max(list(values.values()))[0] - int(tiempoEnvio)]
        print 'Promedio Tiempo de respuesta de las consultas creadas: {}'.format(np.average(list(self.tiempoRespuestaConsultas.values())))
        print 'Varianza Tiempo de respuesta de las consultas creadas: {}'.format(np.var(list(self.tiempoRespuestaConsultas.values())))
        print 'Desviacion estandar Tiempo de respuesta de las consultas creadas: {}'.format(np.std(list(self.tiempoRespuestaConsultas.values())))

    def tiempo_cues_recibidas(self):
        print ':::::::::::::::::::::::::::::::'
        # 1. tiempo-1 de espera de las consultas que arriban
        self.tiempoEsperaConsultasRecibidas = OrderedDict()
        #
        for key, values in self.cuesQueArriban.items():
            for cues in values:
                #print '{}'.format(self.tiempoEsperaConsultas[cues])
                self.tiempoEsperaConsultasRecibidas[cues] = self.tiempoEsperaConsultas[cues]

        print 'Promedio Tiempo de espera de las consultas arribadas: {}'.format(np.average(list(self.tiempoEsperaConsultasRecibidas.values())))
        print 'Varianza Tiempo de espera de las consultas arribadas: {}'.format(np.var(list(self.tiempoEsperaConsultasRecibidas.values())))
        print 'Desviacion estandar de espera de las consultas arribadas: {}'.format(np.std(list(self.tiempoEsperaConsultasRecibidas.values())))

        # 2. Tiempo de servicio de las consultas que arriban
        self.tiempoServicioConsultasRecibidas = OrderedDict()
        for key, values in self.cuesQueArriban.items():
            for cues in values:
                self.tiempoServicioConsultasRecibidas[cues] = self.tiempoServicioConsultas[cues]
        print 'Promedio Tiempo de servicio de las consultas arribadas: {}'.format(np.average(list(self.tiempoServicioConsultasRecibidas.values())))
        print 'Varianza Tiempo de servicio de las consultas arribadas: {}'.format(np.var(list(self.tiempoServicioConsultasRecibidas.values())))
        print 'Desviacion estandar de servicio de las consultas arribadas: {}'.format(np.std(list(self.tiempoServicioConsultasRecibidas.values())))

        # 3. tiempo-1 de respuesta de las consultas que arriban
        self.tiempoRespuestaConsultasArribadas = OrderedDict()
        for key, values in self.cuesQueArriban.items():
            for cues in values:
                self.tiempoRespuestaConsultasArribadas[cues] = self.tiempoRespuestaConsultas[cues]
        print 'Promedio Tiempo de respuesta de las consultas arribadas: {}'.format(np.average(list(self.tiempoRespuestaConsultasArribadas.values())))
        print 'Varianza Tiempo de respuesta de las consultas arribadas: {}'.format(np.var(list(self.tiempoRespuestaConsultasArribadas.values())))
        print 'Desviacion estandar Tiempo de respuesta de las consultas arribadas: {}'.format(np.std(list(self.tiempoRespuestaConsultasArribadas.values())))

    def tiempo_servidor(self):
        print ':::::::::::::::::::::::::::::::'
        #1. tiempo de broadcast
        self.tiempoBroadcastConsultas = OrderedDict()
        self.tiempoBroadcastNoEnviadas = OrderedDict()
        for key, values in self.tiempoElementosCreados.items():
            tiempoarribo = int(str(key).split(':')[2])
            tiempocreacion = int(str(key).split(':')[1])
            if tiempoarribo > 0: # enviadas
                self.tiempoBroadcastConsultas[key] = [tiempoarribo - tiempocreacion]

            else: #no enviada
                self.tiempoBroadcastNoEnviadas[key] = [max(list(values.values()))[0] - tiempocreacion]
        print 'Promedio Tiempo de broadcast de las consultas creadas: {}'.format(np.average(list(self.tiempoBroadcastConsultas.values())))

        if len(list(self.tiempoBroadcastNoEnviadas.values())) == 0:
            self.tiempoBroadcastNoEnviadas['None'] = [0]

        print 'Promedio Tiempo de broadcast de las consultas no enviadas: {}'.format(np.average(list(self.tiempoBroadcastNoEnviadas.values())))

        #2. latencia
        self.latenciaCues = OrderedDict()
        for key, values in self.tiempoElementosCreados.items():
            tiempoarribo = int(str(key).split(':')[2])
            if tiempoarribo > 0: # enviada al servidor
                self.latenciaCues[key] = [min(list(values.values()))[0] - tiempoarribo]
        print 'Promedio latencia de las cues: {}'.format( np.average(list(self.latenciaCues.values())) )



    def cues_creadas_no_enviadas(self):
        print ':::::::::::::::::::::::::::::::'
        self.cuesNoEnviadas = OrderedDict()
        for key, value in self.cuesCreadasPorTiempo.items():
            self.cuesNoEnviadas[key] = []
            for cues in value:
                if int(str(cues).split(':')[2] ) == 0:
                    self.cuesNoEnviadas[key].append(cues)

        for key, value in self.cuesNoEnviadas.items():
            print '{} -> No enviadas: {}'.format(key, len(value))
        print 'Total: {}'.format(sum([len(x) for x in self.cuesNoEnviadas.values()]))

    def talla_cues_creadas(self):
        print ':::::::::::::::::::::::::::::::'
        # 1.1. atributo para almacenar el tamanio de las cues
        self.elementosCuesRespondidas = OrderedDict()
        for  value in self.cuesCreadasRespondidas.values():
            for cue in value:
                self.elementosCuesRespondidas[cue] = []
        #1.2
        self.elementosExactosCuesRespondidas = OrderedDict()
        for value in self.cuesCreadasRespondidas.values():
            for cue in value:
                self.elementosExactosCuesRespondidas[cue] = []
        #2.1
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_ITEM_CUES)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        for r in range(1, max_row):
            for c in range(2, max_column):
                if hojaDeTrabajo.cell(row=r, column=c).value != None:
                    if hojaDeTrabajo.cell(row=r, column=1).value in self.elementosCuesRespondidas:
                        self.elementosCuesRespondidas[hojaDeTrabajo.cell(row=r, column=1).value].append(hojaDeTrabajo.cell(row=r, column=c).value)
        # 2.2
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_ELEMENTOS_EXACTOS_CUES)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        for r in range(1, max_row):
            for c in range(2, max_column):
                if hojaDeTrabajo.cell(row=r, column=c).value != None:
                    if hojaDeTrabajo.cell(row=r, column=1).value in self.elementosExactosCuesRespondidas:
                        self.elementosExactosCuesRespondidas[hojaDeTrabajo.cell(row=r, column=1).value].append(
                            hojaDeTrabajo.cell(row=r, column=c).value)
        #4.1
        self.ElementosCuesRespondidasPorLapsos = OrderedDict()
        for key, value in self.cuesCreadasRespondidas.items():
            self.ElementosCuesRespondidasPorLapsos[key] = OrderedDict()
            for cue in value:
                self.ElementosCuesRespondidasPorLapsos[key][cue] = self.elementosCuesRespondidas[cue]
        # 4.2
        self.ElementosExactosCuesRespondidasPorLapsos = OrderedDict()
        for key, value in self.cuesCreadasRespondidas.items():
            self.ElementosExactosCuesRespondidasPorLapsos[key] = OrderedDict()
            for cue in value:
                self.ElementosExactosCuesRespondidasPorLapsos[key][cue] = self.elementosExactosCuesRespondidas[cue]
        #5.1
        self.tallaCuesRespondidas = OrderedDict()
        for key, value in self.cuesCreadasRespondidas.items():
            self.tallaCuesRespondidas[key] = OrderedDict()
            for cue in value:
                self.tallaCuesRespondidas[key][cue] = len(self.elementosCuesRespondidas[cue])
        #5.2
        self.tallaCuesExactasRespondidas = OrderedDict()
        for key, value in self.cuesCreadasRespondidas.items():
            self.tallaCuesExactasRespondidas[key] = OrderedDict()
            for cue in value:
                self.tallaCuesExactasRespondidas[key][cue] = len(self.elementosExactosCuesRespondidas[cue])

        for key, value in self.ElementosCuesRespondidasPorLapsos.items():
            print key, ' -> itemes de CUEs respondidas: ', len(set().union(*value.values()))
        print ''
        for key, value in self.tallaCuesRespondidas.items():
            print key, ' -> talla Promedio Cues respondidas: ',np.average(list(value.values()))
        print ''
        for key, value in self.tallaCuesExactasRespondidas.items():
            print key, ' -> talla Promedio Cues exactas respondidas: ',np.average(list(value.values()))

    def cues_creadas_por_usuarios(self):
        print ':::::::::::::::::::::::::::::::'
        self.cuesCreadasPorusuarios = OrderedDict()
        for values in self.cuesCreadasPorTiempo.values():
            for cue in values:
                idUsuario = str(cue).split(':')[0]
                if idUsuario in self.cuesCreadasPorusuarios:
                    self.cuesCreadasPorusuarios[idUsuario].append(cue)
                else:
                    self.cuesCreadasPorusuarios[idUsuario] = [cue]

        #for key, values in self.cuesCreadasPorusuarios.items():
            #print key, '-> cues creadas: ', len(values)
        print 'Total de cues credas por usuarios: {}'.format(sum([len(x) for x in self.cuesCreadasPorusuarios.values()]))

    def cues_enviadas_por_usuarios(self):
        print ':::::::::::::::::::::::::::::::'
        self.cuesEnviadasPorUsuarios = OrderedDict()
        for values in self.cuesQueArriban.values():
            for cue in values:
                idUsuario = str(cue).split(':')[0]
                tiempoArribo = int(str(cue).split(':')[2])
                if idUsuario in self.cuesEnviadasPorUsuarios:
                    if tiempoArribo > 0:
                        self.cuesEnviadasPorUsuarios[idUsuario].append(cue)
                else:
                    if tiempoArribo > 0:
                        self.cuesEnviadasPorUsuarios[idUsuario] = [cue]

        #for key, values in self.cuesEnviadasPorUsuarios.items():
            #print key, '-> cues enviadas: ', len(values)
        print 'Total de cues enviadas por usuarios: {}'.format(sum([len(x) for x in self.cuesEnviadasPorUsuarios.values()]))

    def cues_creadas_obsoletas(self):
        print ':::::::::::::::::::::::::::::::'
        # 1.
        cuesRespondidas = []
        for values in self.cuesCreadasPorTiempo.values():
            for cue in values:
                cuesRespondidas.append(cue)
        # 2.
        self.cuesObsoletasCreadas = OrderedDict()
        for key, values in self.cuesCreadasPorusuarios.items():
            self.cuesObsoletasCreadas[key] = []
        #3.
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_CUES_OBSOLETAS)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        for r in range(1, max_row):
            for c in range(2, max_column):
                if hojaDeTrabajo.cell(row=r, column=c).value != None:
                    if hojaDeTrabajo.cell(row=r, column=c).value in cuesRespondidas:
                        idUsuario = str(hojaDeTrabajo.cell(row=r, column=c).value).split(':')[0]
                        self.cuesObsoletasCreadas[idUsuario].append(hojaDeTrabajo.cell(row=r, column=c).value)

        #for key, values in self.cuesObsoletasCreadas.items():
            #print key, '-> obsoletas creadas: ', len(values)
        print 'Total cues creadas obsoletas: {}'.format(sum([len(x) for x in self.cuesObsoletasCreadas.values()]))


    def cues_creadas_obsoletas_enviadas(self):
        print ':::::::::::::::::::::::::::::::'
        # 1.
        cuesRespondidas = []
        for values in self.cuesQueArriban.values():
            for cue in values:
                cuesRespondidas.append(cue)
        # 2.
        self.cuesObsoletasEnviadas = OrderedDict()
        for key, values in self.cuesEnviadasPorUsuarios.items():
            self.cuesObsoletasEnviadas[key] = []
        # 3.
        hojaDeTrabajo = openpyxl.load_workbook(self.URL_EXCEL).get_sheet_by_name(ManejoDeDatos.HOJA_CUES_OBSOLETAS)
        max_row = hojaDeTrabajo.max_row + 1
        max_column = hojaDeTrabajo.max_column + 1
        for r in range(1, max_row):
            for c in range(2, max_column):
                if hojaDeTrabajo.cell(row=r, column=c).value != None:
                    if hojaDeTrabajo.cell(row=r, column=c).value in cuesRespondidas:
                        idUsuario = str(hojaDeTrabajo.cell(row=r, column=c).value).split(':')[0]
                        self.cuesObsoletasEnviadas[idUsuario].append(hojaDeTrabajo.cell(row=r, column=c).value)

        #for key, values in self.cuesObsoletasEnviadas.items():
            #print key, '-> obsoletas creadas: ', len(values)
        print 'Total cues enviadas obsoletas: {}'.format(sum([len(x) for x in self.cuesObsoletasEnviadas.values()]))

    def elementos_usados_por_cronograma(self):
        print ':::::::::::::::::::::::::::::::'
        self.elementosUsadosPorCronogramas = OrderedDict()
        for key, values in self.tallaCuesExactasRespondidas.items():
            self.elementosUsadosPorCronogramas[key] = []
            for cue in values:
                for item in self.elementosExactosCuesRespondidas[cue]:
                    if item not in self.elementosUsadosPorCronogramas[key]:
                        self.elementosUsadosPorCronogramas[key].append(item)
            print '{} -> elementos exactos de cronograma: {}/{}'.format(key, len(self.elementosUsadosPorCronogramas[key]), int(str(key).split(':')[1])-int(str(key).split(':')[0]) )

    def injusticia(self):
        print ':::::::::::::::::::::::::::::::'
        latenciaPromedioCuesArribadas = np.average(list(self.tiempoRespuestaConsultasArribadas.values()))
        sumatoria = 0
        for key, value in self.tiempoRespuestaConsultasArribadas.items():
            print '{} -> {} -> {}'.format(key, value,len(self.tiempoElementosCreados[key]))
            razon = ((value[0] / len(self.tiempoElementosCreados[key])) - latenciaPromedioCuesArribadas)**2
            sumatoria += razon
        print '{} -> {}'.format(sumatoria, len(self.tiempoRespuestaConsultasArribadas))
        sumatoria = sumatoria / len(self.tiempoRespuestaConsultasArribadas)
        print 'Injusticia cues arribadas: {}'.format(sumatoria)

    def razon_de_rendimiento(self):
        print ':::::::::::::::::::::::::::::::'

        keyUltimaCueRespondidas = self.cuesCreadasRespondidas.keys()[self.cuesCreadasRespondidas.__len__()-1]
        #print '{} -> {} -> {}'.format(totalCuesRecibidasRespondidas, keyUltimasCuesRespondidas, len(self.cuesRecibidasRespondidas[keyUltimasCuesRespondidas]))

        self.tiempoEnQueSeRespondeLaUltimaCue = 0
        for cues in self.cuesCreadasRespondidas[keyUltimaCueRespondidas]:
            #print '{} -> {}'.format(cues, max(list(self.tiempoElementosCreados[cues].values()))[0])
            # se busca la cue que recibe el utlimo elemento de dato
            if max(list(self.tiempoElementosCreados[cues].values()))[0] > self.tiempoEnQueSeRespondeLaUltimaCue :
                self.tiempoEnQueSeRespondeLaUltimaCue = max(list(self.tiempoElementosCreados[cues].values()))[0]
        print 'razon de rendimiento: {} / {}'.format( self.cuesTotalesCreadas, self.tiempoEnQueSeRespondeLaUltimaCue)

    def utilidad_y_efectividad_del_programa(self):
        print ':::::::::::::::::::::::::::::::'
        self.cuesCreadasConElementosEnCronograma = OrderedDict()
        self.utilidadDeProgramas = OrderedDict()
        self.efectividadProgramas = OrderedDict()
        for key, value in self.programas.items():
            self.cuesCreadasConElementosEnCronograma[str(key)] = []
            self.utilidadDeProgramas[str(key)] = []
            self.efectividadProgramas[str(key)] = []
            intervaloInicio = int(str(key).split(':')[0])
            for k, v in self.cuesCreadasPorTiempo.items():
                if int(k)+1 == intervaloInicio:
                    self.cuesCreadasConElementosEnCronograma[str(key)] = v


        for key, value in self.cuesCreadasConElementosEnCronograma.items():
            #print '{} -> {}'.format(key, len(value))
            intervaloFin = int(str(key).split(':')[1])
            cuesEnIntervalo = []
            elementosEnIntervalo = []
            for cue in value:
                #if cue in  self.tiempoElementosCreados:
                for elemento, tiempo in self.tiempoElementosCreados[cue].items():
                    if tiempo[0] < intervaloFin:
                        #print '   {}-{}->{}'.format(cue, elemento, tiempo)
                        if cue not in cuesEnIntervalo:
                            cuesEnIntervalo.append(cue)
                        if elemento not in elementosEnIntervalo:
                            elementosEnIntervalo.append(elemento)

            #self.cuesCreadasConElementosEnCronograma[key] = cuesEnIntervalo
            #print '{} -> {}'.format(key, len(cuesEnIntervalo))
            itemessolicitados = []
            for cue in cuesEnIntervalo:
                for item in self.elementosExactosCuesRespondidas[cue]:
                    if item not in itemessolicitados:
                        itemessolicitados.append(item)

            print '{}'.format(key)
            if len(elementosEnIntervalo) > 0:
                self.utilidadDeProgramas[key].append(Decimal(len(elementosEnIntervalo))/Decimal(int(str(key).split(':')[1])-int(str(key).split(':')[0])))
                self.utilidadDeProgramas[key].append(Decimal(len(elementosEnIntervalo)))
                self.utilidadDeProgramas[key].append(Decimal(int(str(key).split(':')[1])-int(str(key).split(':')[0])))

            print '  Utilidad: {} / {}'.format( len(elementosEnIntervalo), int(str(key).split(':')[1])-int(str(key).split(':')[0]))
            if len(elementosEnIntervalo)>0:
                self.efectividadProgramas[key].append(Decimal(len(elementosEnIntervalo))/ Decimal( len(itemessolicitados)))
                self.efectividadProgramas[key].append(Decimal(len(elementosEnIntervalo)))
                self.efectividadProgramas[key].append(Decimal( len(itemessolicitados)))
            #
            print '  Efectividad: {} / {}'.format(len(elementosEnIntervalo), len(itemessolicitados))
            #os.system('pause')
        self.utilidadDeProgramas = {k: v for k, v in self.utilidadDeProgramas.items() if len(v) > 0}
        self.utilidadDeProgramas = OrderedDict(sorted(self.utilidadDeProgramas.items(), key=lambda x: int(str(x[0]).split(':')[0])))
        self.efectividadProgramas = {k: v for k, v in self.efectividadProgramas.items() if len(v) > 0}
        self.efectividadProgramas = OrderedDict(sorted(self.efectividadProgramas.items(), key=lambda x: int(str(x[0]).split(':')[0])))

    def justicia(self):
        print ':::::::::::::::::::::::::::::::'
        self.jitterPromedioPorPrograma = OrderedDict()
        self.stretchPromedioPorPrograma = OrderedDict()
        for cronograma, cues in self.cuesCreadasRespondidas.items():
            distanciasPrograma = []
            stretchPrograma = []
            for cue in cues:
                tiempos = []
                distancias = []
                stretch = []
                if len(self.tiempoElementosCreados[cue].values()) >1:
                    for tiempo in self.tiempoElementosCreados[cue].values():
                        tiempos.append(tiempo[0])
                        if len(tiempos)> 1:
                            distancias.append(tiempos[len(tiempos)-1] - tiempos[len(tiempos)-2])
                    tiempoEspera = max(tiempos)- min(tiempos)
                    tiempoServicio = len(tiempos)
                    stretch.append( float(tiempoEspera)/tiempoServicio )
                if len(self.tiempoElementosCreados[cue].values()) == 1:
                    distancias.append(0)
                    stretch.append(0)
                #print '{}'.format(distancias)
                #for elemento, tiempo in self.tiempoElementosCreados[cue].items():
                    #print '        {}->{}'.format(elemento, tiempo)
                #print '{}'.format(tiempos)
                #if cronograma == '5068:5422':
                    #print '    {}->distancia promedio: {} --> {}'.format(cue, np.average(distancias), distancias)
                distanciasPrograma.append(np.nanmean(distancias))
                stretchPrograma.append(np.nanmean(stretch))
            print '{}->Jitter:{}->Stretch:{}'.format(cronograma, np.nanmean(distanciasPrograma), np.nanmean(stretchPrograma))
            self.jitterPromedioPorPrograma[cronograma] = np.nanmean(distanciasPrograma)
            self.stretchPromedioPorPrograma[cronograma] = np.nanmean(stretchPrograma)





def cambiar_direccion_pruebas(k, e, t):
    os.chdir('..')
    direccionPruebas = "[K={}-R={}-T={}]".format(k, e, t)
    os.chdir(direccionPruebas)
    return os.getcwd()

def crear_carpeta(path):
    access_rights = 0o755
    try:
        os.mkdir(path, access_rights)
    except OSError:
        print ("Creation of the directory %s failed" % path)
    else:
        print ("Successfully created the directory %s" % path)

def graficar(path, datos, seleccion):
    #Grafico.cues_encubiertas_enviadas(path, datos)
    #Grafico.cues_creadas(path, datos)
    #Grafico.razon_de_privacidad(path, datos)
    #Grafico.razon_de_rendimiento(path, datos)
    Grafico.cues_creadas_respondidas(path, datos)
    #Grafico.cues_recibidas_respondidas(path, datos)
    #Grafico.razon_utilidad(path, datos)
    #Grafico.razon_efectividad(path, datos)
    #Grafico.razon_cues_arribadas_respondidas(path, datos)
    #Grafico.tiempo_servidor(path, datos)
    #Grafico.tiempo_servidor_no_enviadas(path, datos)
    #Grafico.tiempo_servidor_total(path, datos)
    #Grafico.obsolescencia(path, datos)
    #Grafico.puntos_seleccion(path, datos, seleccion)
    #Grafico.tabla_carga_trabajo(path, datos)
    #Grafico.tabla_tiempo_respuesta(path, datos)
    #Grafico.justicia_stetch(path, datos)
    #Grafico.justicia_jitter(path, datos)

if __name__ == '__main__':

    Estadisticas.URL_EXCEL = 'D:\Usuarios\PABLO\Proyectos\Python\workspaces-2019-1\schedule\src\manejoDeDatos\[K=4-R=3-T=50]\[AlgoritmoDeEnvergaduraProbabilista][Jitter][USER=200][POI=1600][k=4].xlsx'
    e1 = Estadisticas(1600, 200, 4, 3, 50, "AEP-ASJ-T-50")
    Estadisticas.URL_EXCEL = 'D:\Usuarios\PABLO\Proyectos\Python\workspaces-2019-1\schedule\src\manejoDeDatos\[K=4-R=3-T=50]\[AlgoritmoDeEnvergaduraProbabilista][Stretch][USER=200][POI=1600][k=4].xlsx'
    e2 = Estadisticas(1600, 200, 4, 3, 50, "AEP-ASS-T-50")
    Estadisticas.URL_EXCEL = 'D:\Usuarios\PABLO\Proyectos\Python\workspaces-2019-1\schedule\src\manejoDeDatos\[K=4-R=3-T=50]\[AlgoritmoDeEnvergaduraProbabilista][CargaDeTrabajo][USER=200][POI=1600][k=4].xlsx'
    e3 = Estadisticas(1600, 200, 4, 3, 50, "AEP-ASC-T-50")
    graficar('D:\Usuarios\PABLO\Proyectos\Python\workspaces-2019-1\schedule\src\manejoDeDatos\graficos', [e1, e2, e3], False)
    os.system('pause')

    algoritmosScheduleDeterminista = [
                                      ['\[AlgoritmoDeEnvergaduraDeterminista]', 'AED'],#0
                                      ['\[AlgoritmoDePopularidadDeterminista]', 'APD'],#1
                                      ['\[AlgoritmoRelevanciaDeterminista]', 'ARD']#2
                                     ]
    algoritmosScheduleProbabilista =[
                                      ['\[AlgoritmoDeEnvergaduraProbabilista]', 'AEP'],#0
                                      ['\[AlgoritmoDePopularidadProbabilista]', 'APP'],#1
                                      ['\[AlgoritmoRelevanciaProbabilista]', 'ARP'],#2
                                      ['\[AlgoritmoDeEnvergaduraProbabilistaConTiempoDeEspera]', 'AEPT'],#3
                                      ['\[AlgoritmoDePopularidadProbabilistaConTiempoDeEspera]', 'APPT'],#4
                                      ['\[AlgoritmoRelevanciaProbabilistaConTiempoDeEspera]', 'ARPT']#5
                                    ]
    algoritmoSeleccion = [
                          ['[CargaDeTrabajo]', 'ASC'],#0
                          ['[Jitter]', 'ASJ'],#1
                          ['[Stretch]', 'ASS']#0
                          ]
    # -------------------------------------
    os.chdir('graficos')
    direccionDeGraficos = os.getcwd()
    # -------------------------------------
    direccionPruebas = cambiar_direccion_pruebas(4,3,100)
    # -------------------------------------
    numeroCarpeta = 1
    # -------------------------------------

    print ''
    print '::::::::::::::::::::'
    print '   VARIANDO USUARIOS'
    print '::::::::::::::::::::'
    #[1-3]
    for asd in algoritmosScheduleDeterminista:
        print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
        d1 = direccionPruebas+asd[0]+"[None]"+"[USER={}][POI={}][k={}].xlsx".format(200, 1600, 4)
        d2 = direccionPruebas+asd[0]+"[None]"+"[USER={}][POI={}][k={}].xlsx".format(300, 1600, 4)
        d3 = direccionPruebas+asd[0]+"[None]"+"[USER={}][POI={}][k={}].xlsx".format(400, 1600, 4)
        d4 = direccionPruebas+asd[0]+"[None]"+"[USER={}][POI={}][k={}].xlsx".format(500, 1600, 4)
        #-------------------------------------
        Estadisticas.URL_EXCEL = d1
        e1 = Estadisticas(1600, 200, 4, 3, 100, "{}-{}".format(asd[1], 200))
        Estadisticas.URL_EXCEL = d2
        e2 = Estadisticas(1600, 300, 4, 3, 100, "{}-{}".format(asd[1], 300))
        Estadisticas.URL_EXCEL = d3
        e3 = Estadisticas(1600, 400, 4, 3, 100, "{}-{}".format(asd[1], 400))
        Estadisticas.URL_EXCEL = d4
        e4 = Estadisticas(1600, 500, 4, 3, 100, "{}-{}".format(asd[1], 500))
        #---------------------------------------
        datos = [e1, e2, e3, e4]
        carpetaDatos = '\datos-{}'.format(numeroCarpeta)
        crear_carpeta(direccionDeGraficos + carpetaDatos)
        #---------------------------------------
        graficar(direccionDeGraficos+carpetaDatos, datos, False)

        print d1
        print d2
        print d3
        print d4
        numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print '   VARIANDO USUARIOS'
    print '::::::::::::::::::::'
    #[4-21]
    for asd in algoritmosScheduleProbabilista:
        for ads in algoritmoSeleccion:
            print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
            d1 = direccionPruebas + asd[0] +ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 4)
            d2 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(300, 1600, 4)
            d3 = direccionPruebas + asd[0] +ads[0] + "[USER={}][POI={}][k={}].xlsx".format(400, 1600, 4)
            d4 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(500, 1600, 4)
            # -------------------------------------
            Estadisticas.URL_EXCEL = d1
            e1 = Estadisticas(1600, 200, 4, 3, 100, "{}-{}-{}".format(asd[1], ads[1], 200))
            Estadisticas.URL_EXCEL = d2
            e2 = Estadisticas(1600, 300, 4, 3, 100, "{}-{}-{}".format(asd[1], ads[1], 300))
            Estadisticas.URL_EXCEL = d3
            e3 = Estadisticas(1600, 400, 4, 3, 100, "{}-{}-{}".format(asd[1], ads[1], 400))
            Estadisticas.URL_EXCEL = d4
            e4 = Estadisticas(1600, 500, 4, 3, 100, "{}-{}-{}".format(asd[1], ads[1], 500))
            # ---------------------------------------
            datos = [e1, e2, e3, e4]
            carpetaDatos = '\datos-{}'.format(numeroCarpeta)
            crear_carpeta(direccionDeGraficos + carpetaDatos)
            # ---------------------------------------
            graficar(direccionDeGraficos + carpetaDatos, datos, True)
            print d1
            print d2
            print d3
            print d4
            numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print '  VARIANDO ALGORITMO'
    print '::::::::::::::::::::'
    #[22-25]
    for users in [200, 300, 400, 500]:
        print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
        d1 = direccionPruebas + algoritmosScheduleDeterminista[0][0] + "[None]" + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
        d2 = direccionPruebas + algoritmosScheduleDeterminista[1][0] + "[None]" + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
        d3 = direccionPruebas + algoritmosScheduleDeterminista[2 ][0] + "[None]" + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
        # -------------------------------------
        Estadisticas.URL_EXCEL = d1
        e1 = Estadisticas(1600, users, 4, 3, 100, "{}".format(algoritmosScheduleDeterminista[0][1]))
        Estadisticas.URL_EXCEL = d2
        e2 = Estadisticas(1600, users, 4, 3, 100, "{}".format(algoritmosScheduleDeterminista[1][1]))
        Estadisticas.URL_EXCEL = d3
        e3 = Estadisticas(1600, users, 4, 3, 100, "{}".format(algoritmosScheduleDeterminista[2][1]))
        # ---------------------------------------
        datos = [e1, e2, e3]
        carpetaDatos = '\datos-{}'.format(numeroCarpeta)
        crear_carpeta(direccionDeGraficos + carpetaDatos)
        # ---------------------------------------
        graficar(direccionDeGraficos+carpetaDatos, datos, False)
        print d1
        print d2
        print d3
        numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print '  VARIANDO ALGORITMO'
    print '::::::::::::::::::::'
    #[26-37]
    for users in [200, 300, 400, 500]:
        for ads in algoritmoSeleccion:
            print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
            d1 = direccionPruebas + algoritmosScheduleProbabilista[0][0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
            d2 = direccionPruebas + algoritmosScheduleProbabilista[1][0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
            d3 = direccionPruebas + algoritmosScheduleProbabilista[2][0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
            # -------------------------------------
            Estadisticas.URL_EXCEL = d1
            e1 = Estadisticas(1600, users, 4, 3, 100, "{}-{}".format(algoritmosScheduleProbabilista[0][1],ads[1]))
            Estadisticas.URL_EXCEL = d2
            e2 = Estadisticas(1600, users, 4, 3, 100, "{}-{}".format(algoritmosScheduleProbabilista[1][1],ads[1]))
            Estadisticas.URL_EXCEL = d3
            e3 = Estadisticas(1600, users, 4, 3, 100, "{}-{}".format(algoritmosScheduleProbabilista[2][1],ads[1]))
            # ---------------------------------------
            datos = [e1, e2, e3]
            carpetaDatos = '\datos-{}'.format(numeroCarpeta)
            crear_carpeta(direccionDeGraficos + carpetaDatos)
            # ---------------------------------------
            graficar(direccionDeGraficos+carpetaDatos, datos, True)
            print d1
            print d2
            print d3
            numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print '  VARIANDO ALGORITMO'
    print '::::::::::::::::::::'
    #[38-49]
    for users in [200, 300, 400, 500]:
        for ads in algoritmoSeleccion:
            print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
            d1 = direccionPruebas + algoritmosScheduleProbabilista[3][0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
            d2 = direccionPruebas + algoritmosScheduleProbabilista[4][0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
            d3 = direccionPruebas + algoritmosScheduleProbabilista[5][0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(users, 1600, 4)
            # -------------------------------------
            Estadisticas.URL_EXCEL = d1
            e1 = Estadisticas(1600, users, 4, 3, 100, "{}-{}".format(algoritmosScheduleProbabilista[3][1],ads[1]))
            Estadisticas.URL_EXCEL = d2
            e2 = Estadisticas(1600, users, 4, 3, 100, "{}-{}".format(algoritmosScheduleProbabilista[4][1],ads[1]))
            Estadisticas.URL_EXCEL = d3
            e3 = Estadisticas(1600, users, 4, 3, 100, "{}-{}".format(algoritmosScheduleProbabilista[5][1],ads[1]))
            # ---------------------------------------
            datos = [e1, e2, e3]
            carpetaDatos = '\datos-{}'.format(numeroCarpeta)
            crear_carpeta(direccionDeGraficos + carpetaDatos)
            # ---------------------------------------
            graficar(direccionDeGraficos+carpetaDatos, datos, True)
            print d1
            print d2
            print d3
            numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print 'VARIANDO K ANONIMATO'
    print '::::::::::::::::::::'
    #[50-52]L
    for asd in algoritmosScheduleDeterminista:
        for ads in ['[None]']:
            print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
            direccionPruebas = cambiar_direccion_pruebas(4, 3, 100)
            d1 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 4)
            direccionPruebas = cambiar_direccion_pruebas(6, 3, 100)
            d2 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 6)
            direccionPruebas = cambiar_direccion_pruebas(8, 3, 100)
            d3 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 8)
            direccionPruebas = cambiar_direccion_pruebas(10, 3, 100)
            d4 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 10)
            # -------------------------------------
            Estadisticas.URL_EXCEL = d1
            e1 = Estadisticas(1600, 200, 4, 3, 100, "{}-{}".format(asd[1], "K-4"))
            Estadisticas.URL_EXCEL = d2
            e2 = Estadisticas(1600, 200, 6, 3, 100, "{}-{}".format(asd[1], "K-6"))
            Estadisticas.URL_EXCEL = d3
            e3 = Estadisticas(1600, 200, 8, 3, 100, "{}-{}".format(asd[1], "K-8"))
            Estadisticas.URL_EXCEL = d4
            e4 = Estadisticas(1600, 200, 10, 3, 100, "{}-{}".format(asd[1], "K-10"))
            # ---------------------------------------
            datos = [e1, e2, e3, e4]
            carpetaDatos = '\datos-{}'.format(numeroCarpeta)
            crear_carpeta(direccionDeGraficos + carpetaDatos)
            # ---------------------------------------
            graficar(direccionDeGraficos+carpetaDatos, datos, False)
            print d1
            print d2
            print d3
            print d4
            numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print 'VARIANDO K ANONIMATO'
    print '::::::::::::::::::::'
    #[53-70]L
    for asd in algoritmosScheduleProbabilista:
        for ads in algoritmoSeleccion:
            print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
            direccionPruebas = cambiar_direccion_pruebas(4, 3, 100)
            d1 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 4)
            direccionPruebas = cambiar_direccion_pruebas(6, 3, 100)
            d2 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 6)
            direccionPruebas = cambiar_direccion_pruebas(8, 3, 100)
            d3 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 8)
            direccionPruebas = cambiar_direccion_pruebas(10, 3, 100)
            d4 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, 10)
            # -------------------------------------
            Estadisticas.URL_EXCEL = d1
            e1 = Estadisticas(1600, 200, 4, 3, 100, "{}-{}-{}".format(asd[1], ads[1], "K-4"))
            Estadisticas.URL_EXCEL = d2
            e2 = Estadisticas(1600, 200, 6, 3, 100, "{}-{}-{}".format(asd[1], ads[1], "K-6"))
            Estadisticas.URL_EXCEL = d3
            e3 = Estadisticas(1600, 200, 8, 3, 100, "{}-{}-{}".format(asd[1], ads[1], "K-8"))
            Estadisticas.URL_EXCEL = d4
            e4 = Estadisticas(1600, 200, 10, 3, 100, "{}-{}-{}".format(asd[1], ads[1], "K-10"))
            # ---------------------------------------
            datos = [e1, e2, e3, e4]
            carpetaDatos = '\datos-{}'.format(numeroCarpeta)
            crear_carpeta(direccionDeGraficos + carpetaDatos)
            # ---------------------------------------
            graficar(direccionDeGraficos+carpetaDatos, datos, True)
            print d1
            print d2
            print d3
            print d4
            numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print 'VARIANDO EFECTIVIDAD'
    print '::::::::::::::::::::'
    #[71-82]
    for asd in algoritmosScheduleDeterminista:
        for anmt in [4,6,8,10]:
            for ads in ['[None]']:
                print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
                direccionPruebas = cambiar_direccion_pruebas(anmt, 1, 100)
                d1 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 100)
                d2 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 6, 100)
                d3 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                # -------------------------------------
                Estadisticas.URL_EXCEL = d1
                e1 = Estadisticas(1600, 200, anmt, 1, 100, "{}-{}".format(asd[1], "R-1"))
                Estadisticas.URL_EXCEL = d2
                e2 = Estadisticas(1600, 200, anmt, 3, 100, "{}-{}".format(asd[1], "R-3"))
                Estadisticas.URL_EXCEL = d3
                e3 = Estadisticas(1600, 200, anmt, 6, 100, "{}-{}".format(asd[1], "R-6"))
                # ---------------------------------------
                datos = [e1, e2, e3]
                carpetaDatos = '\datos-{}'.format(numeroCarpeta)
                crear_carpeta(direccionDeGraficos + carpetaDatos)
                # ---------------------------------------
                graficar(direccionDeGraficos+carpetaDatos, datos, False)
                print d1
                print d2
                print d3
                numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print 'VARIANDO EFECTIVIDAD'
    print '::::::::::::::::::::'
    #[83-154]
    for asd in algoritmosScheduleProbabilista:
        for anmt in [4, 6, 8, 10]:
            for ads in algoritmoSeleccion:
                print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
                direccionPruebas = cambiar_direccion_pruebas(anmt, 1, 100)
                d1 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 100)
                d2 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 6, 100)
                d3 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                # -------------------------------------
                Estadisticas.URL_EXCEL = d1
                e1 = Estadisticas(1600, 200, anmt, 1, 100, "{}-{}-{}".format(asd[1], ads[1], "R-1"))
                Estadisticas.URL_EXCEL = d2
                e2 = Estadisticas(1600, 200, anmt, 3, 100, "{}-{}-{}".format(asd[1], ads[1], "R-3"))
                Estadisticas.URL_EXCEL = d3
                e3 = Estadisticas(1600, 200, anmt, 6, 100, "{}-{}-{}".format(asd[1], ads[1], "R-6"))
                # ---------------------------------------
                datos = [e1, e2, e3]
                carpetaDatos = '\datos-{}'.format(numeroCarpeta)
                crear_carpeta(direccionDeGraficos + carpetaDatos)
                # ---------------------------------------
                graficar(direccionDeGraficos+carpetaDatos, datos, True)
                print d1
                print d2
                print d3
                numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print '      VARIANDO TALLA'
    print '::::::::::::::::::::'
    #[155-166]
    for asd in algoritmosScheduleDeterminista:
        for anmt in [4, 6, 8, 10]:
            for ads in ['[None]']:
                print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 50)
                d1 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 100)
                d2 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 200)
                d3 = direccionPruebas + asd[0] + ads + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                # -------------------------------------
                Estadisticas.URL_EXCEL = d1
                e1 = Estadisticas(1600, 200, anmt, 3, 50, "{}-{}".format(asd[1], "T-50"))
                Estadisticas.URL_EXCEL = d2
                e2 = Estadisticas(1600, 200, anmt, 3, 100, "{}-{}".format(asd[1], "T-100"))
                Estadisticas.URL_EXCEL = d3
                e3 = Estadisticas(1600, 200, anmt, 3, 200, "{}-{}".format(asd[1], "T-200"))
                # ---------------------------------------
                datos = [e1, e2, e3]
                carpetaDatos = '\datos-{}'.format(numeroCarpeta)
                crear_carpeta(direccionDeGraficos + carpetaDatos)

                # ---------------------------------------
                graficar(direccionDeGraficos+carpetaDatos, datos, False)
                print d1
                print d2
                print d3
                numeroCarpeta += 1

    print ''
    print '::::::::::::::::::::'
    print '      VARIANDO TALLA'
    print '::::::::::::::::::::'
    #[167-238]
    for asd in algoritmosScheduleProbabilista:
        for anmt in [4, 6, 8, 10]:
            for ads in algoritmoSeleccion:
                print ':::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::'
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 50)
                d1 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 100)
                d2 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                direccionPruebas = cambiar_direccion_pruebas(anmt, 3, 200)
                d3 = direccionPruebas + asd[0] + ads[0] + "[USER={}][POI={}][k={}].xlsx".format(200, 1600, anmt)
                # -------------------------------------
                Estadisticas.URL_EXCEL = d1
                e1 = Estadisticas(1600, 200, anmt, 3, 50, "{}-{}-{}".format(asd[1], ads[1], "T-50"))
                Estadisticas.URL_EXCEL = d2
                e2 = Estadisticas(1600, 200, anmt, 3, 100, "{}-{}-{}".format(asd[1], ads[1], "T-100"))
                Estadisticas.URL_EXCEL = d3
                e3 = Estadisticas(1600, 200, anmt, 3, 200, "{}-{}-{}".format(asd[1], ads[1], "T-200"))
                # ---------------------------------------
                datos = [e1, e2, e3]
                carpetaDatos = '\datos-{}'.format(numeroCarpeta)
                crear_carpeta(direccionDeGraficos + carpetaDatos)
                # ---------------------------------------
                graficar(direccionDeGraficos+carpetaDatos, datos, True)
                print d1
                print d2
                print d3
                numeroCarpeta += 1