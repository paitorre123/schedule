import xlsxwriter
import openpyxl
import os
import math
from src.consulta.consultaDeRango import ConsultaDeRango
from src.consulta.consultaDeRangoArtificial import ConsultaDeRangoArtificial


class ManejoDeDatos(object):

    NOMBRE_ARCHIVO = ''
    PATH_ARCHIVO = '../manejoDeDatos/'
    HOJA_SERVER_CUES = 'SERVER CUES'
    HOJA_USER_CREATED_CUES = 'USER CREATED CUES'
    HOJA_ITEM_CUES = 'ITEM CUES'
    HOJA_ITEM_SUBCONSULTAS = 'ITEM SUBCONSULTAS'
    HOJA_CUES_RESPONDIDAS = 'CUES RESPONDIDAS'
    HOJA_TIEMPO_ESPERA_CUES_EXACTAS = 'TIEMPO ESPERA CUES EXACTAS'
    HOJA_TIEMPO_ESPERA_CUES_VIRTUALES= 'TIEMPO ESPERA CUES VIRTUALES'
    HOJA_ELEMENTOS_EXACTOS_CUES = 'ELEMENTOS EXACTOS CUES'
    HOJA_ELEMENTOS_VIRTUALES_CUES = 'ELEMENTOS VIRTUALES CUES'
    HOJA_CUES_USAN_ELEMENTOS = 'CUES QUE USAN ELEMENTOS'
    HOJA_CUES_OBSOLETAS = 'CUES OBSOLETAS'
    HOJA_VALORES_SELECCION = 'VALORES SELECCION'
    _extension = '.xlsx'

    @classmethod
    def _selecionar_archivo(cls):
        if os.path.exists(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension):
            return openpyxl.load_workbook(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        else:
            # 1- crear archivo
            archivoDeTrabajo = xlsxwriter.Workbook(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
            # 2- crear hojas
            archivoDeTrabajo.add_worksheet(cls.HOJA_SERVER_CUES)
            archivoDeTrabajo.add_worksheet(cls.HOJA_USER_CREATED_CUES)
            archivoDeTrabajo.add_worksheet(cls.HOJA_ITEM_CUES)
            archivoDeTrabajo.add_worksheet(cls.HOJA_ITEM_SUBCONSULTAS)
            archivoDeTrabajo.add_worksheet(cls.HOJA_CUES_RESPONDIDAS)
            archivoDeTrabajo.add_worksheet(cls.HOJA_TIEMPO_ESPERA_CUES_EXACTAS)
            archivoDeTrabajo.add_worksheet(cls.HOJA_TIEMPO_ESPERA_CUES_VIRTUALES)
            archivoDeTrabajo.add_worksheet(cls.HOJA_ELEMENTOS_EXACTOS_CUES)
            archivoDeTrabajo.add_worksheet(cls.HOJA_ELEMENTOS_VIRTUALES_CUES)
            archivoDeTrabajo.add_worksheet(cls.HOJA_CUES_USAN_ELEMENTOS)
            archivoDeTrabajo.add_worksheet(cls.HOJA_CUES_OBSOLETAS)
            archivoDeTrabajo.add_worksheet(cls.HOJA_VALORES_SELECCION)
            #3- cerrar el archivo
            archivoDeTrabajo.close()
            #
            #
            return openpyxl.load_workbook(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)

    @classmethod
    def escribir_valor_seleccion_cronograma(cls, valorSeleccion, largo):

            # 1- abrir archivo de trabajo
            archivoDeTrabajo_escritura = cls._selecionar_archivo()
            # 2- selecionar hoja correspondiente
            hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_VALORES_SELECCION)
            # 3.ver filas actuales
            max_row_hojaServerCues = hojaServerCues.max_row + 1
            # 4- ingresar los datos a la hoja
            col = 2
            hojaServerCues.cell(row=max_row_hojaServerCues, column=1, value='Cronograma:{}'.format(max_row_hojaServerCues))
            hojaServerCues.cell(row=max_row_hojaServerCues, column=col, value=valorSeleccion)
            col = 3
            hojaServerCues.cell(row=max_row_hojaServerCues, column=col, value=largo)
            # 5- guardar y cerrar el archivo
            archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
            archivoDeTrabajo_escritura.close()

    @classmethod
    def escribir_server_cues(cls, cues, time):
        if len(cues) > 0:
            #1- abrir archivo de trabajo
            archivoDeTrabajo_escritura = cls._selecionar_archivo()
            #2- selecionar hoja correspondiente
            hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_SERVER_CUES)
            #3.ver filas actuales
            max_row_hojaServerCues = hojaServerCues.max_row + 1
            #4- ingresar los datos a la hoja
            col = 2
            hojaServerCues.cell(row=max_row_hojaServerCues,column=1,value=time)
            for cue in cues:
                hojaServerCues.cell(row=max_row_hojaServerCues,column=col,value='{}'.format(cue))
                col += 1
            #5- guardar y cerrar el archivo
            archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
            archivoDeTrabajo_escritura.close()

    @classmethod
    def escribir_user_created_cues(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_USER_CREATED_CUES)

        # 3- ingresar los datos a la hoja
        row = 1
        col = 1
        for user in users:
            hojaServerCues.cell(row=row, column=col, value='{}'.format(user))
            col += 1
            for q in user.almacenes:
                hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                col += 1
            col = 1
            row += 1

        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def escribir_item_cues(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_ITEM_CUES)

        # 3- ingresar los datos a la hoja
        row = 1
        col = 1
        for user in users:
            for q in user.almacenes:
                hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                col += 1
                for e in q.elementosRequeridosReales:
                    hojaServerCues.cell(row=row, column=col, value='{}'.format(e))
                    col += 1
                for e in q.elementosRequeridosArtificiales:
                    hojaServerCues.cell(row=row, column=col, value='{}'.format(e))
                    col += 1
                row += 1
                col = 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def escribir_item_subconsultas(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_ITEM_SUBCONSULTAS)

        # 3- ingresar los datos a la hoja
        row = 1
        col = 1
        for user in users:
            for a in user.almacenes:
                for sq in a.subConsultasDeRango:
                    hojaServerCues.cell(row=row, column=col, value='{}'.format(sq))
                    col += 1
                    for e in sq.elementosRequeridos:
                        hojaServerCues.cell(row=row, column=col, value='{}'.format(e))
                        col += 1
                    row += 1
                    col = 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def numero_cues_respondidas(cls, users, historialDeCronogramas):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_CUES_RESPONDIDAS)
        #2.1- calcular el numero de cues respondidas

        time = 1
        for subCronograma in historialDeCronogramas:
            # print 'Subcronograma: {}'.format(nSubcronograma)

            row = 1
            for elemento in subCronograma.items:
                if elemento.__class__.__name__ is not 'ElementoFinBroadcast':
                    #print 'time:{}'.format(time)
                    cuesRespondidas = cls._CUESQueResponde(users, time)
                    col = 1
                    #print 'time:{}'.format(time)
                    hojaServerCues.cell(row=time, column=col, value=time)
                    for cr in cuesRespondidas:
                        col += 1
                        hojaServerCues.cell(row=time, column=col, value='{}'.format(cr))
                time += 1

        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def _CUESQueResponde(cls, users, time):
        respondidas = []
        for user in users:
            for q in user.almacenes:
                if len(q.elementosEncontradosReales) > 0:
                    tiempoUltimoElementoReal = q.tiempoElementosEncontradosReales[len(q.tiempoElementosEncontradosReales) - 1]
                    if time == tiempoUltimoElementoReal:
                        respondidas.append(q.consultaEncubierta)
        return respondidas

    @classmethod
    def tiempo_de_espera_cues_exactas(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_TIEMPO_ESPERA_CUES_EXACTAS)
        #3- guardar tiempo-1 de espera de las cues exactas
        row = 1
        for user in users:
            for q in user.almacenes:
                col = 1
                # print 'time:{}'.format(time)
                hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                col += 1
                if len(q.elementosEncontradosReales) > 0:
                    tiempoPrimerElementoReal = q.tiempoElementosEncontradosReales[0]
                    tiempoUltimoElementoReal = q.tiempoElementosEncontradosReales[len(q.tiempoElementosEncontradosReales) - 1]
                    hojaServerCues.cell(row=row, column=col, value=tiempoUltimoElementoReal-tiempoPrimerElementoReal)
                row += 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def tiempo_de_espera_cues_virtuales(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_TIEMPO_ESPERA_CUES_VIRTUALES)
        # 3- guardar tiempo-1 de espera de las cues exactas
        row = 1
        for user in users:
            for q in user.almacenes:
                col = 1
                # print 'time:{}'.format(time)
                hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                col += 1
                if len(q.elementosEncontradosArtificiales) > 0:
                    tiempoPrimerElementoArtificial = q.tiempoElementosEncontradosArtificiales[0]
                    tiempoUltimoElementoArtificial = q.tiempoElementosEncontradosArtificiales[len(q.tiempoElementosEncontradosArtificiales) - 1]
                    hojaServerCues.cell(row=row, column=col, value=tiempoUltimoElementoArtificial - tiempoPrimerElementoArtificial)
                row += 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def numero_elementos_exactos_cues(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_ELEMENTOS_EXACTOS_CUES)
        # 3- guardar tiempo-1 de espera de las cues exactas
        row = 1
        for user in users:
            for q in user.almacenes:
                col = 1
                # print 'time:{}'.format(time)
                hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                col += 1
                if len(q.elementosRequeridosReales) > 0:
                    for er in q.elementosRequeridosReales:
                        hojaServerCues.cell(row=row, column=col, value='{}'.format(er))
                        col += 1
                row += 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def numero_elementos_virtuales_cues(cls, users):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_ELEMENTOS_VIRTUALES_CUES)
        # 3- guardar tiempo-1 de espera de las cues exactas
        row = 1
        for user in users:
            for q in user.almacenes:
                col = 1
                # print 'time:{}'.format(time)
                hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                col += 1
                if len(q.elementosRequeridosArtificiales) > 0:
                    for ev in q.elementosRequeridosArtificiales:
                        hojaServerCues.cell(row=row, column=col, value='{}'.format(ev))
                        col += 1
                row += 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def cues_exactas_que_usan_elementos(cls, users, historialDeCronogramas):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_CUES_USAN_ELEMENTOS)
        # 3- guardar tiempo-1 de espera de las cues exactas
        time = 1
        for subCronograma in historialDeCronogramas:
            # print 'Subcronograma: {}'.format(nSubcronograma)

            for elemento in subCronograma.items:
                if elemento.__class__.__name__ is not 'ElementoFinBroadcast':
                    # print 'time:{}'.format(time)
                    cuesQueUsan = cls._CUESExactasQueLoUsan(users, time)
                    col = 1
                    # print 'time:{}'.format(time)
                    hojaServerCues.cell(row=time, column=col, value='{}'.format(elemento.puntoDeInteres))
                    for cqu in cuesQueUsan:
                        col += 1
                        hojaServerCues.cell(row=time, column=col, value='{}'.format(cqu))
                time += 1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

    @classmethod
    def _CUESExactasQueLoUsan(self, users, time):
        cues = []
        for user in users:
            for q in user.almacenes:
                if time in q.tiempoElementosEncontradosReales:
                    cues.append(q.consultaEncubierta)
        return cues

    @classmethod
    def cues_obsoletas(cls, users, grid):
        # 1- abrir archivo de trabajo
        archivoDeTrabajo_escritura = cls._selecionar_archivo()
        # 2- selecionar hoja correspondiente
        hojaServerCues = archivoDeTrabajo_escritura.get_sheet_by_name(cls.HOJA_CUES_OBSOLETAS)
        row = 1
        for user in users:
            col =1
            hojaServerCues.cell(row=row, column=col, value='{}'.format(user))
            col += 1
            for q in user.almacenes:
                if len(q.elementosEncontradosReales) > 0:
                    tiempoPrimerElementoReal = int(q.tiempoElementosEncontradosReales[0])
                    tiempoUltimoElementoReal = int(q.tiempoElementosEncontradosReales[
                        len(q.tiempoElementosEncontradosReales) - 1])
                    xInicial = user.randomWalk.obtener_psicion_en_tiempo(tiempoPrimerElementoReal-1)[0]
                    yInicial = user.randomWalk.obtener_psicion_en_tiempo(tiempoPrimerElementoReal-1)[1]
                    xFinal = user.randomWalk.obtener_psicion_en_tiempo(tiempoUltimoElementoReal-1)[0]
                    yFinal = user.randomWalk.obtener_psicion_en_tiempo(tiempoUltimoElementoReal-1)[1]
                    distancia = math.sqrt( math.pow((xFinal-xInicial), 2) + math.pow((yFinal-yInicial), 2) )
                    celdaFinal = grid.cellOfUser(xFinal, yFinal)
                    if distancia > celdaFinal.side*3:
                        hojaServerCues.cell(row=row, column=col, value='{}'.format(q.consultaEncubierta))
                        col += 1
            row +=1
        # 4- guardar y cerrar el archivo
        archivoDeTrabajo_escritura.save(cls.PATH_ARCHIVO + cls.NOMBRE_ARCHIVO + cls._extension)
        archivoDeTrabajo_escritura.close()

