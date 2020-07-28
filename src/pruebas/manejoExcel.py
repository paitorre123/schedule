import xlsxwriter
import openpyxl
import os
from src.client.cliente import Usuario
from collections import OrderedDict
import re


class ManejoExcel(object):
    fileName = 'None'
    _extension = '.xlsx'
    urlSave = '../pruebas/datos/'
    urlRead = 'None'
    @classmethod
    def crear_archivo_excel(cls):
        archivoDeTrabajo = xlsxwriter.Workbook(cls.urlSave+cls.fileName+cls._extension)
        return archivoDeTrabajo

    @classmethod
    def crear_hoja_en_archivo(cls, archivoDeTrabajo, nombreHoja):
        hoja = archivoDeTrabajo.add_worksheet(nombreHoja)
        return hoja

    @classmethod
    def escribir_datos_usuarios(cls, archivoDeTrabajo, usuarios):
        #archivoDeTrabajo.write(row, col, time)
        row = 0
        col = 0
        #x = []
        #y = []
        for user in usuarios:
            #print 'user position: {},{}'.format(user.point.pointX, user.point.pointY)
            for position in user.randomWalk.rastro:
                #print '{}'.format(position)
                archivoDeTrabajo.write(row, col, '{},{}'.format(position[0], position[1]))
                row+=1
            col+=1
            row=0

    @classmethod
    def escribir_datos_poi(cls, archivoDeTrabajo, grid):
        row = 0
        col = 0
        for c in grid.cells:
            for poi in c.pointsInterest:
                print '{},{}'.format(poi.point.pointX,poi.point.pointX)
                archivoDeTrabajo.write(row, col, '{},{}'.format(poi.point.pointX,poi.point.pointY))
                row+=1

    @classmethod
    def escribir_datos_cues(cls, archivoDeTrabajo, row , col, cue):
        archivoDeTrabajo.write(row, col, cue)


    @classmethod
    def abrir_archivo_excel(cls):
        return openpyxl.load_workbook(cls.urlRead)

    @classmethod
    def leer_datos_usuarios(cls, archivoDeTrabajo, nombreHoja):
        hojaTrabajo = archivoDeTrabajo.get_sheet_by_name(nombreHoja)

        # get max row count
        max_row = hojaTrabajo.max_row
        # get max column count
        max_column = hojaTrabajo.max_column
        # iterate over all cells
        # iterate over all rows
        usuariosRastro = []
        for i in range(1, max_column + 1):
            # iterate over all columns
            rastro = []
            for j in range(1, max_row + 1):
                # get particular cell value
                cell_obj = hojaTrabajo.cell(row=j, column=i)
                # print cell value
                if cell_obj.value != None:
                    x,y = cell_obj.value.split(',')
                    x=float(x)
                    y=float(y)
                    #print'x:{},y:{}'.format(x,y)
                    rastro.append([x,y,j])
             # print new line
            #print('\n')
            usuariosRastro.append(rastro)

        return usuariosRastro

    @classmethod
    def leer_anonimato_usuarios(cls, archivoDeTrabajo, nombreHoja):
        hojaTrabajo = archivoDeTrabajo.get_sheet_by_name(nombreHoja)

        # get max row count
        max_row = hojaTrabajo.max_row
        # get max column count
        max_column = hojaTrabajo.max_column
        # iterate over all cells
        # iterate over all rows
        usuariosAnonimato = []
        for i in range(1, max_row + 1):
            # iterate over all columns
            anonimato = []
            for j in range(1, max_column + 1):
                # get particular cell value
                cell_obj = hojaTrabajo.cell(row=i, column=j)
                # print cell value
                if cell_obj.value != None:
                    k = cell_obj.value
                    k = int(k)
                    # print'x:{},y:{}'.format(x,y)
                    anonimato.append(k)
            # print new line
            # print('\n')
            usuariosAnonimato.append(anonimato)

        return usuariosAnonimato

    @classmethod
    def leer_celdas_cues_usuarios(cls, archivoDeTrabajo, nombreHoja):
        hojaTrabajo = archivoDeTrabajo.get_sheet_by_name(nombreHoja)

        # get max row count
        max_row = hojaTrabajo.max_row
        # get max column count
        max_column = hojaTrabajo.max_column
        # iterate over all cells
        # iterate over all rows
        usuariosceldasCues = OrderedDict()
        for i in range(1, max_row + 1):
            # iterate over all columns
            usuariosceldasCues[i] = []

            for j in range(1, max_column + 1):
                # get particular cell value
                cell_obj = hojaTrabajo.cell(row=i, column=j)
                # print cell value
                celdas = map(int,list(filter(None, str(cell_obj.value).split(':') )))
                usuariosceldasCues[i].append(celdas)

            # print('\n')


        return usuariosceldasCues





